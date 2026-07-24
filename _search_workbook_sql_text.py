import openpyxl

from config import WORKBOOK_PATH

wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False, keep_vba=True)
needles = ("select", "from", "where", "join", "tempdb", "#zone", "industry")

for ws in wb.worksheets:
    found = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            v = cell.value
            if isinstance(v, str):
                s = v.lower()
                if any(n in s for n in needles):
                    if found == 0:
                        print(f"\n[{ws.title}]")
                    found += 1
                    print(f"{cell.coordinate}: {v[:200].replace(chr(10), ' ')}")
    if found:
        print(f"-- matches: {found}")
