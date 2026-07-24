from pathlib import Path
import openpyxl

from config import WORKBOOK_PATH

wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False, keep_vba=True)
ws = wb["SQL"]

for r in range(1, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 12) + 1)]
    if any(v not in (None, "") for v in vals):
        print(r, "|", " | ".join(str(v) if v is not None else "" for v in vals))
